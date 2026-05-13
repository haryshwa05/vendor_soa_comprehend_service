from __future__ import annotations

import csv
from dataclasses import dataclass
from datetime import datetime
from html import unescape
from io import BytesIO
from pathlib import Path
import re
from typing import Iterable

import extract_msg
import pdfplumber
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Border, Font, PatternFill, Side


class MsgConversionError(Exception):
    """Raised when a .msg file cannot be converted into a worksheet."""


@dataclass
class ExtractedTable:
    headers: list[str]
    rows: list[list[str]]


@dataclass
class CanonicalRow:
    invoice_number: str
    invoice_date: str
    outstanding_amount: str


@dataclass
class ExtractionResult:
    rows: list[CanonicalRow]
    source_type: str
    source_name: str
    source_details: str
    output_path: Path | None = None


CANONICAL_HEADERS = ["INV Number", "INV Date", "Outstanding Amount"]
SUPPORTED_EXCEL_SUFFIXES = {".xlsx", ".xlsm", ".csv"}
HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(color="FFFFFF", bold=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="D1D5DB"),
    right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"),
    bottom=Side(style="thin", color="D1D5DB"),
)
INVOICE_NUMBER_ALIASES = (
    "invoice number",
    "invoice no",
    "invoice #",
    "invoice",
    "inv number",
    "inv no",
    "inv #",
    "inv#",
    "document no",
    "document number",
)
INVOICE_DATE_ALIASES = (
    "invoice date",
    "inv date",
    "date",
    "bill date",
)
AMOUNT_ALIASES = (
    "outstanding amount",
    "outstanding",
    "open balance",
    "balance due",
    "balance",
    "amount due",
    "total due",
    "amount",
)


def convert_msg_to_excel(msg_path: str | Path, output_dir: str | Path | None = None) -> Path:
    result = extract_msg_to_rows(msg_path)
    output_path = next_available_output_path(Path(msg_path), output_dir=output_dir)
    write_rows_to_workbook(result.rows, output_path)
    result.output_path = output_path
    return output_path


def extract_msg_to_rows(msg_path: str | Path) -> ExtractionResult:
    msg_file = Path(msg_path).expanduser().resolve()
    validate_msg_path(msg_file)
    message = open_message(msg_file)

    extractors = (
        extract_rows_from_message_body,
        extract_rows_from_excel_attachments,
        extract_rows_from_pdf_attachments,
    )
    for extractor in extractors:
        result = extractor(message)
        if result.rows:
            return result

    raise MsgConversionError(
        "No usable invoice details were found in the email body or supported attachments."
    )


def workbook_bytes_from_msg(msg_path: str | Path) -> tuple[ExtractionResult, bytes]:
    result = extract_msg_to_rows(msg_path)
    return result, workbook_bytes_from_rows(result.rows)


def workbook_bytes_from_rows(rows: list[CanonicalRow]) -> bytes:
    workbook = build_workbook(rows)
    buffer = BytesIO()
    workbook.save(buffer)
    workbook.close()
    return buffer.getvalue()


def rows_to_dicts(rows: list[CanonicalRow]) -> list[dict[str, str]]:
    return [
        {
            "invoice_number": row.invoice_number,
            "invoice_date": row.invoice_date,
            "outstanding_amount": row.outstanding_amount,
        }
        for row in rows
    ]


def validate_msg_path(msg_path: Path) -> None:
    if not msg_path.exists():
        raise MsgConversionError(f"File not found: {msg_path}")
    if msg_path.suffix.lower() != ".msg":
        raise MsgConversionError("Only .msg files are supported.")


def open_message(msg_path: Path):
    try:
        return extract_msg.Message(str(msg_path))
    except Exception as exc:
        raise MsgConversionError(f"Unable to open .msg file: {exc}") from exc


def extract_rows_from_message_body(message) -> ExtractionResult:
    html_body = getattr(message, "htmlBody", None)
    if not html_body:
        return empty_result("body", "email_body", "Email did not contain an HTML body.")

    best_rows: list[CanonicalRow] = []
    best_name = "html_table"
    for index, table in enumerate(extract_html_tables(html_body), start=1):
        rows = canonicalize_table(table)
        if len(rows) > len(best_rows):
            best_rows = rows
            best_name = f"html_table_{index}"

    if not best_rows:
        return empty_result(
            "body",
            "email_body",
            "Email body contained HTML but no table mapped to invoice/date/amount.",
        )
    return ExtractionResult(
        rows=best_rows,
        source_type="body",
        source_name=best_name,
        source_details="Extracted from the largest usable HTML table in the email body.",
    )


def extract_rows_from_excel_attachments(message) -> ExtractionResult:
    best_rows: list[CanonicalRow] = []
    best_name = ""
    best_details = ""

    for attachment in getattr(message, "attachments", []) or []:
        filename = get_attachment_filename(attachment)
        suffix = Path(filename).suffix.lower()
        if suffix not in SUPPORTED_EXCEL_SUFFIXES:
            continue

        file_bytes = getattr(attachment, "data", None)
        if not isinstance(file_bytes, (bytes, bytearray)) or not file_bytes:
            continue

        try:
            if suffix in {".xlsx", ".xlsm"}:
                rows, detail = extract_rows_from_excel_bytes(bytes(file_bytes))
            else:
                rows, detail = extract_rows_from_csv_bytes(bytes(file_bytes))
        except Exception:
            continue

        if len(rows) > len(best_rows):
            best_rows = rows
            best_name = filename
            best_details = detail

    if not best_rows:
        return empty_result(
            "excel_attachment",
            "attachment_scan",
            "No supported Excel attachment produced mapped invoice rows.",
        )
    return ExtractionResult(
        rows=best_rows,
        source_type="excel_attachment",
        source_name=best_name,
        source_details=best_details,
    )


def extract_rows_from_pdf_attachments(message) -> ExtractionResult:
    best_rows: list[CanonicalRow] = []
    best_name = ""
    best_details = ""
    for attachment in getattr(message, "attachments", []) or []:
        filename = get_attachment_filename(attachment)
        if not filename.lower().endswith(".pdf"):
            continue

        pdf_bytes = getattr(attachment, "data", None)
        if not isinstance(pdf_bytes, (bytes, bytearray)) or not pdf_bytes:
            continue

        try:
            rows, detail = extract_rows_from_pdf_bytes(bytes(pdf_bytes))
        except Exception:
            continue

        if len(rows) > len(best_rows):
            best_rows = rows
            best_name = filename
            best_details = detail

    if not best_rows:
        return empty_result(
            "pdf_attachment",
            "attachment_scan",
            "No PDF attachment produced mapped invoice rows.",
        )
    return ExtractionResult(
        rows=best_rows,
        source_type="pdf_attachment",
        source_name=best_name,
        source_details=best_details,
    )


def empty_result(source_type: str, source_name: str, source_details: str) -> ExtractionResult:
    return ExtractionResult(
        rows=[],
        source_type=source_type,
        source_name=source_name,
        source_details=source_details,
    )


def get_attachment_filename(attachment) -> str:
    for attr in ("longFilename", "shortFilename", "displayName", "name"):
        value = getattr(attachment, attr, None)
        if value:
            return str(value)
    try:
        return str(attachment.getFilename())
    except Exception:
        return "attachment"


def extract_html_tables(html_body: bytes | str) -> list[ExtractedTable]:
    soup = BeautifulSoup(html_body, "html.parser")
    tables = soup.find_all("table")
    parsed_tables: list[ExtractedTable] = []
    for table in tables:
        parsed = parse_html_table(table)
        if parsed and parsed.rows:
            parsed_tables.append(parsed)
    return parsed_tables


def extract_rows_from_pdf_bytes(pdf_bytes: bytes) -> tuple[list[CanonicalRow], str]:
    best_rows: list[CanonicalRow] = []
    best_details = "No usable PDF rows found."
    with pdfplumber.open(BytesIO(pdf_bytes)) as pdf:
        for page_number, page in enumerate(pdf.pages, start=1):
            for table_index, raw_table in enumerate(page.extract_tables() or [], start=1):
                table = extracted_table_from_matrix(raw_table)
                rows = canonicalize_table(table)
                if len(rows) > len(best_rows):
                    best_rows = rows
                    best_details = f"PDF table extraction from page {page_number}, table {table_index}."

            text = page.extract_text() or ""
            rows = extract_rows_from_pdf_text(text)
            if len(rows) > len(best_rows):
                best_rows = rows
                best_details = f"PDF text fallback extraction from page {page_number}."
    return best_rows, best_details


def extract_rows_from_excel_bytes(file_bytes: bytes) -> tuple[list[CanonicalRow], str]:
    workbook = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    best_rows: list[CanonicalRow] = []
    best_details = "No usable worksheet rows found."
    try:
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            matrix: list[list[str]] = []
            for row in worksheet.iter_rows(values_only=True):
                cleaned = [clean_cell_text(value) for value in row]
                if any(cell for cell in cleaned):
                    matrix.append(cleaned)
            table = extracted_table_from_matrix(matrix)
            rows = canonicalize_table(table)
            if len(rows) > len(best_rows):
                best_rows = rows
                best_details = f"Excel attachment extraction from worksheet '{sheet_name}'."
    finally:
        workbook.close()
    return best_rows, best_details


def extract_rows_from_csv_bytes(file_bytes: bytes) -> tuple[list[CanonicalRow], str]:
    text = decode_text_bytes(file_bytes)
    reader = csv.reader(text.splitlines())
    matrix: list[list[str]] = []
    for row in reader:
        cleaned = [clean_cell_text(value) for value in row]
        if any(cell for cell in cleaned):
            matrix.append(cleaned)
    table = extracted_table_from_matrix(matrix)
    return canonicalize_table(table), "CSV attachment extraction."


def decode_text_bytes(file_bytes: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            return file_bytes.decode(encoding)
        except UnicodeDecodeError:
            continue
    return file_bytes.decode("utf-8", errors="replace")


def extracted_table_from_matrix(raw_table: list[list[str | None]]) -> ExtractedTable | None:
    normalized = []
    for row in raw_table:
        cleaned = [clean_cell_text(cell) for cell in row]
        if any(cell for cell in cleaned):
            normalized.append(cleaned)

    if len(normalized) < 2:
        return None

    width = max(len(row) for row in normalized)
    normalized_rows = [pad_row(row, width) for row in normalized]
    header_index = find_header_row(normalized_rows)
    if header_index is None or header_index == len(normalized_rows) - 1:
        return None

    headers = dedupe_headers(normalized_rows[header_index])
    rows = [row for row in normalized_rows[header_index + 1 :] if any(cell.strip() for cell in row)]
    if not rows:
        return None
    return ExtractedTable(headers=headers, rows=rows)


def parse_html_table(table) -> ExtractedTable | None:
    grid: list[list[str]] = []
    span_map: dict[tuple[int, int], str] = {}
    row_index = 0

    for tr in table.find_all("tr"):
        cells = tr.find_all(["th", "td"])
        if not cells:
            continue

        row: list[str] = []
        col_index = 0

        while (row_index, col_index) in span_map:
            row.append(span_map.pop((row_index, col_index)))
            col_index += 1

        for cell in cells:
            while (row_index, col_index) in span_map:
                row.append(span_map.pop((row_index, col_index)))
                col_index += 1

            value = clean_cell_text(cell.get_text(" ", strip=True))
            colspan = parse_span(cell.get("colspan"))
            rowspan = parse_span(cell.get("rowspan"))

            for _ in range(colspan):
                row.append(value)
                if rowspan > 1:
                    for offset in range(1, rowspan):
                        span_map[(row_index + offset, col_index)] = value
                col_index += 1

        while (row_index, col_index) in span_map:
            row.append(span_map.pop((row_index, col_index)))
            col_index += 1

        if any(cell.strip() for cell in row):
            grid.append(row)
        row_index += 1

    if not grid:
        return None

    width = max(len(row) for row in grid)
    normalized_rows = [pad_row(row, width) for row in grid]
    header_index = find_header_row(normalized_rows)
    if header_index is None or header_index == len(normalized_rows) - 1:
        return None

    headers = dedupe_headers(normalized_rows[header_index])
    data_rows = [
        row
        for row in normalized_rows[header_index + 1 :]
        if any(cell.strip() for cell in row)
    ]
    if not data_rows:
        return None

    return ExtractedTable(headers=headers, rows=data_rows)


def parse_span(value: str | None) -> int:
    if not value:
        return 1
    try:
        return max(int(value), 1)
    except ValueError:
        return 1


def pad_row(row: list[str], width: int) -> list[str]:
    if len(row) >= width:
        return row[:]
    return row + [""] * (width - len(row))


def find_header_row(rows: list[list[str]]) -> int | None:
    best_index = None
    best_score = -1
    for idx, row in enumerate(rows[:-1]):
        non_empty = [cell for cell in row if cell.strip()]
        if len(non_empty) < 2:
            continue

        text_like = sum(1 for cell in non_empty if has_letters(cell))
        unique_ratio = len(set(cell.lower() for cell in non_empty)) / max(len(non_empty), 1)
        score = (text_like * 5) + int(unique_ratio * 10)

        if score > best_score:
            best_score = score
            best_index = idx

    return best_index


def has_letters(value: str) -> bool:
    return any(char.isalpha() for char in value)


def dedupe_headers(headers: Iterable[str]) -> list[str]:
    seen: dict[str, int] = {}
    cleaned_headers: list[str] = []

    for idx, header in enumerate(headers, start=1):
        base = clean_header_name(header, idx)
        count = seen.get(base, 0) + 1
        seen[base] = count
        cleaned_headers.append(base if count == 1 else f"{base}_{count}")

    return cleaned_headers


def clean_header_name(value: str, position: int) -> str:
    cleaned = clean_cell_text(value)
    if not cleaned:
        return f"Column_{position}"
    return re.sub(r"\s+", " ", cleaned).strip()


def clean_cell_text(value: str | None) -> str:
    if value is None:
        return ""
    text = unescape(str(value)).replace("\xa0", " ")
    text = re.sub(r"\s+", " ", text).strip()
    return text


def next_available_output_path(msg_path: Path, output_dir: str | Path | None = None) -> Path:
    target_dir = Path(output_dir).expanduser().resolve() if output_dir else Path(__file__).resolve().parent
    candidate = target_dir / f"{msg_path.stem}_vendor_soa.xlsx"
    if not candidate.exists():
        return candidate

    stem = f"{msg_path.stem}_vendor_soa"
    counter = 1
    while True:
        candidate = target_dir / f"{stem}_{counter}.xlsx"
        if not candidate.exists():
            return candidate
        counter += 1


def write_rows_to_workbook(rows: list[CanonicalRow], output_path: Path) -> None:
    workbook = build_workbook(rows)
    workbook.save(output_path)
    workbook.close()


def build_workbook(rows: list[CanonicalRow]) -> Workbook:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Vendor SOA"

    sheet.append(CANONICAL_HEADERS)
    for row in rows:
        sheet.append([row.invoice_number, row.invoice_date, row.outstanding_amount])

    for cell in sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.border = THIN_BORDER

    autosize_columns(sheet)
    return workbook


def autosize_columns(sheet) -> None:
    for column_cells in sheet.columns:
        letter = column_cells[0].column_letter
        width = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[letter].width = min(max(width + 2, 10), 45)


def canonicalize_table(table: ExtractedTable | None) -> list[CanonicalRow]:
    if table is None:
        return []
    mapping = resolve_canonical_mapping(table.headers)
    if mapping.get("invoice_number") is None or mapping.get("outstanding_amount") is None:
        return []

    results: list[CanonicalRow] = []
    for row in table.rows:
        invoice_number = normalize_invoice_number(get_row_value(row, mapping.get("invoice_number")))
        invoice_date = normalize_date(get_row_value(row, mapping.get("invoice_date")))
        outstanding_amount = normalize_amount(get_row_value(row, mapping.get("outstanding_amount")))

        if not invoice_number or not outstanding_amount:
            continue

        results.append(
            CanonicalRow(
                invoice_number=invoice_number,
                invoice_date=invoice_date,
                outstanding_amount=outstanding_amount,
            )
        )
    return dedupe_canonical_rows(results)


def resolve_canonical_mapping(headers: list[str]) -> dict[str, int | None]:
    normalized_headers = [normalize_header(header) for header in headers]
    return {
        "invoice_number": find_best_header(normalized_headers, INVOICE_NUMBER_ALIASES),
        "invoice_date": find_best_header(normalized_headers, INVOICE_DATE_ALIASES),
        "outstanding_amount": find_best_header(normalized_headers, AMOUNT_ALIASES),
    }


def normalize_header(value: str) -> str:
    value = clean_cell_text(value).lower()
    value = re.sub(r"[^a-z0-9]+", " ", value)
    return re.sub(r"\s+", " ", value).strip()


def find_best_header(headers: list[str], aliases: Iterable[str]) -> int | None:
    best_index = None
    best_score = -1
    for index, header in enumerate(headers):
        if not header:
            continue
        for alias in aliases:
            score = score_header_alias(header, alias)
            if score > best_score:
                best_score = score
                best_index = index
    if best_score < 50:
        return None
    return best_index


def score_header_alias(header: str, alias: str) -> int:
    alias_normalized = normalize_header(alias)
    if header == alias_normalized:
        return 100
    if alias_normalized in header or header in alias_normalized:
        return 90

    header_tokens = set(header.split())
    alias_tokens = set(alias_normalized.split())
    if not alias_tokens:
        return 0
    overlap = len(header_tokens & alias_tokens)
    return int((overlap / len(alias_tokens)) * 100)


def get_row_value(row: list[str], index: int | None) -> str:
    if index is None or index >= len(row):
        return ""
    return clean_cell_text(row[index])


def normalize_invoice_number(value: str) -> str:
    value = clean_cell_text(value)
    if not value:
        return ""
    value = re.sub(r"\.0$", "", value)
    return value


def normalize_amount(value: str) -> str:
    value = clean_cell_text(value)
    if not value:
        return ""
    value = value.replace(",", "")
    value = re.sub(r"^\((.*)\)$", r"-\1", value)
    match = re.search(r"-?\$?\d+(?:\.\d+)?", value)
    if not match:
        return ""
    number = match.group(0).replace("$", "")
    try:
        return f"{float(number):.2f}"
    except ValueError:
        return ""


def normalize_date(value: str) -> str:
    value = clean_cell_text(value)
    if not value:
        return ""
    for pattern in ("%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d", "%d/%m/%Y", "%d-%b-%Y", "%d-%B-%Y"):
        try:
            return datetime.strptime(value, pattern).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return value


def extract_rows_from_pdf_text(text: str) -> list[CanonicalRow]:
    rows: list[CanonicalRow] = []
    for line in text.splitlines():
        cleaned = clean_cell_text(line)
        if not cleaned:
            continue
        if "invoice" in cleaned.lower() and "balance" in cleaned.lower():
            continue

        date_match = re.search(r"\b(?:\d{1,2}/\d{1,2}/\d{2,4}|\d{4}-\d{2}-\d{2})\b", cleaned)
        amount_matches = re.findall(r"\$?\d[\d,]*\.\d{2}\b", cleaned)
        invoice_number = invoice_number_from_text_line(cleaned, date_match.group(0) if date_match else "")

        if not date_match or not amount_matches or not invoice_number:
            continue

        rows.append(
            CanonicalRow(
                invoice_number=normalize_invoice_number(invoice_number),
                invoice_date=normalize_date(date_match.group(0)),
                outstanding_amount=normalize_amount(amount_matches[-1]),
            )
        )
    return dedupe_canonical_rows([row for row in rows if row.invoice_number and row.outstanding_amount])


def dedupe_canonical_rows(rows: list[CanonicalRow]) -> list[CanonicalRow]:
    seen: set[tuple[str, str, str]] = set()
    deduped: list[CanonicalRow] = []
    for row in rows:
        key = (row.invoice_number, row.invoice_date, row.outstanding_amount)
        if key in seen:
            continue
        seen.add(key)
        deduped.append(row)
    return deduped


def invoice_number_from_text_line(line: str, date_value: str) -> str:
    if not date_value:
        return ""

    prefix = line.split(date_value, 1)[0].strip()
    if not prefix:
        return ""

    tokens = re.findall(r"[A-Z0-9][A-Z0-9\-\/]*", prefix, flags=re.IGNORECASE)
    if not tokens:
        return ""

    preferred = [token for token in tokens if any(char.isdigit() for char in token)]
    if preferred:
        return preferred[-1]
    return tokens[-1]
